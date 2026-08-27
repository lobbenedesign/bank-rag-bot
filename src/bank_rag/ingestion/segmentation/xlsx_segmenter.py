"""Segments an Excel workbook per sheet, in fixed-size row batches. Locator
value is "SheetName!first-last" so a no-index rule can exclude e.g. rows on
one pricing tab without touching other sheets in the same workbook.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from bank_rag.domain.entities import ChunkLocator, DocumentSegment

_ROWS_PER_SEGMENT = 50


def segment_xlsx(content: bytes) -> list[DocumentSegment]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    segments: list[DocumentSegment] = []

    for sheet in workbook.worksheets:
        rows = [
            [str(cell) if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)
        ]
        if not rows:
            continue
        header, data_rows = rows[0], rows[1:]
        for start in range(0, len(data_rows), _ROWS_PER_SEGMENT):
            batch = data_rows[start : start + _ROWS_PER_SEGMENT]
            lines = [", ".join(header)] + [", ".join(row) for row in batch]
            first, last = start + 2, start + 1 + len(batch)
            segments.append(
                DocumentSegment(
                    text="\n".join(lines),
                    locator=ChunkLocator(kind="row_range", value=f"{sheet.title}!{first}-{last}"),
                )
            )
    return segments
